import openpyxl, csv, os
from openpyxl.utils import get_column_letter

for excelFile in os.listdir(os.getcwd()):
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if len(wb.sheetnames) == 1:
                outFile = open('{}.csv'.format(excelFile.replace('.xlsx','_new')),'w', newline='')
            else:
                outFile = open('{}_{}.csv'.format(excelFile.replace('.xlsx',''), sheetname), 'w', newline= '')
            outWriter = csv.writer(outFile)

            for rowNum in range(sheet.max_row):
                row = []
                for colNum in range(sheet.max_column):
                    cell = sheet[get_column_letter(colNum+1) + str(rowNum+1)].value
                    row.append(cell)
                outWriter.writerow(row)
            outFile.close()
        wb.close()