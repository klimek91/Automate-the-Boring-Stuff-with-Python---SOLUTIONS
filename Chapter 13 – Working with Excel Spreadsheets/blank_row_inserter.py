#!python

import openpyxl, sys
from openpyxl.utils import get_column_letter

if len(sys.argv) > 3:
    file = sys.argv[3]
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    n = sys.argv[1]
    m = sys.argv[2]

    for row in range(sheet.max_row+m,n,-1):
        for col in range(1,sheet.max_column+1):
            sheet[get_column_letter(col) + str(row)].value = sheet[get_column_letter(col) + str(row-m)].value

    for row in range(n,n+m):
        for col in range(1,sheet.max_column+1):
            del sheet[get_column_letter(col) + str(row)]

    wb.save('new'+file)