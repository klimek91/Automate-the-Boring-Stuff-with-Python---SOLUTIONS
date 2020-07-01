import openpyxl
from openpyxl.utils import get_column_letter

file = input("Enter xlsx file name which You want to invert: ")

wb = openpyxl.load_workbook(file)
sheet = wb.active

cols = []
for row in range(1, sheet.max_row+1):
    rows = []
    for col in range(1, sheet.max_column+1):
        rows.append(sheet[get_column_letter(col)+str(row)].value)
    cols.append(rows)

print(cols)
wb = openpyxl.Workbook()
sheet = wb.active

for col in range(len(cols)):
    for row in range(len(cols[0])):
        sheet[get_column_letter(col+1) + str(row+1)].value = cols[col][row]

wb.save('inverted_'+file)