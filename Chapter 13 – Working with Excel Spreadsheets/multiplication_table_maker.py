#!python

import openpyxl, sys
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

size = 10
for col in range(1, size+1):
    sheet['A' + str(col + 1)].value = col
    sheet[get_column_letter(col + 1) + '1'].value = col

for col in range(2, size+2):
    for row in range(2, size+2):
        sheet[get_column_letter(col) + str(row)].value = sheet[get_column_letter(col) + str(1)].value * sheet['A'+str(row)].value

wb.save('multiplication_table.xlsx')