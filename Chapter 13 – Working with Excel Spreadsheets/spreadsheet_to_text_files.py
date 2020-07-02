import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('textToExcel.xlsx')
sheet = wb.active

for row in range(sheet.max_row):
    for col in range(sheet.max_column):
        file = open('text{}.txt'.format(col), 'a')
        file.write(str(sheet[get_column_letter(col+1) + str(row+1)].value))






#for row in range(sheet.max_row):
#    file = open('textFromExcel01.txt', 'a')
#    file.write(sheet['A'+str(row+1)].value)
