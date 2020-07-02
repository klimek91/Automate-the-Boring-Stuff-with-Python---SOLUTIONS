import openpyxl, os
from openpyxl.utils import get_column_letter

files = []  # text files in current dir
for file in os.listdir(os.getcwd()):
    if file.endswith('.txt'):
        files.append(file)

wb = openpyxl.Workbook()
sheet = wb.active

for file in files:
    text = open(file)
    lines = text.readlines()
    for row in range(len(lines)):
        sheet[get_column_letter(files.index(file)+1)+str(row+1)].value = lines[row]
        if sheet.column_dimensions[get_column_letter(files.index(file)+1)].width < len(lines[row]):
            sheet.column_dimensions[get_column_letter(files.index(file) + 1)].width = len(lines[row])

wb.save("textToExcel.xlsx")