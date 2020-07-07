import smtplib, openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.active

members = {}

for col in range(1,sheet.max_column+1):
    for row in range(1,sheet.max_row):
        if sheet[get_column_letter(col) + str(row)].value == None:
            members.setdefault(sheet['A'+str(row)].value, {})
            members[sheet['A'+str(row)].value].setdefault(sheet['B'+str(row)].value, [])
            members[sheet['A' + str(row)].value][sheet['B'+str(row)].value].append(sheet[get_column_letter(col) + "1"].value)

email = input("Please enter Your email address: ")
password = input("Please enter Your password: ")

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(email, password)

for member in members:
    for email,months in members[member].items():
        smtpObj.sendmail(email, 'Subject: Reminder of unpaid membership.\n\
        Dear {}, You have unpaid dues for {}, please make payment asap.'.format(member, ', '.join(months)))

smtpObj.quit()